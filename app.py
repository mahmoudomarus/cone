#!/usr/bin/env python3
"""
Invoice Scanner Web App
Upload multiple invoices and get a combined Excel file
"""

from flask import Flask, render_template, request, send_file, flash, redirect, url_for
import os
import base64
import json
from datetime import datetime
from dotenv import load_dotenv
from werkzeug.utils import secure_filename
import tempfile
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from PIL import Image
import io
import google.generativeai as genai

# Load environment variables
load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your-secret-key-change-this')
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024  # 20MB max upload (reduced)
app.config['MAX_FILES'] = 10  # Limit to 10 files at once

# Get Google API key
API_KEY = os.getenv("GOOGLE_API_KEY")
if not API_KEY:
    raise Exception("Please set GOOGLE_API_KEY in .env file")

# Configure Gemini
genai.configure(api_key=API_KEY)
model = genai.GenerativeModel('gemini-1.5-flash')

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def compress_image(image_path, max_size=(1920, 1920), quality=85):
    """Compress image to reduce memory usage"""
    try:
        img = Image.open(image_path)
        
        # Convert to RGB if needed
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        
        # Resize if too large
        img.thumbnail(max_size, Image.Resampling.LANCZOS)
        
        # Save compressed version
        output = io.BytesIO()
        img.save(output, format='JPEG', quality=quality, optimize=True)
        output.seek(0)
        
        return output.getvalue()
    except Exception as e:
        print(f"Compression error: {e}, using original")
        with open(image_path, 'rb') as f:
            return f.read()

def encode_image(image_path):
    """Encode image to base64 with compression"""
    compressed_data = compress_image(image_path)
    return base64.b64encode(compressed_data).decode('utf-8')

def extract_invoice_data(image_path):
    """Use Google Gemini Vision API to extract invoice data in structured format"""
    
    prompt = """Extract ALL line items from this invoice/receipt and flatten them into a simple list.

Return JSON in this EXACT format:
{
  "date": "采购时间：2020.10.1",
  "items": [
    {"品名": "海带丝", "数量": "1", "单价": "5.00", "金额": "5.00"},
    {"品名": "大头菜(颗)", "数量": "2.1", "单价": "1.70", "金额": "3.57"},
    {"品名": "土豆", "数量": "2.2", "单价": "1.28", "金额": "2.82"}
  ]
}

RULES:
- Extract EVERY product/item from the invoice
- If invoice has multiple columns, extract ALL items from ALL columns into one flat list
- Keep exact Chinese text for product names (品名)
- 数量 = quantity 
- 单价 = unit price
- 金额 = total amount
- Return ONLY valid JSON, no markdown, no explanation, no code blocks"""

    try:
        # Open and send image to Gemini
        img = Image.open(image_path)
        
        response = model.generate_content([prompt, img])
        content = response.text.strip()
        
        # Clean up response
        if content.startswith("```json"):
            content = content[7:]
        if content.startswith("```"):
            content = content[3:]
        if content.endswith("```"):
            content = content[:-3]
        
        content = content.strip()
        
        print(f"Gemini response: {content[:200]}...")
        
        return json.loads(content)
    
    except Exception as e:
        print(f"Gemini API error: {e}")
        import traceback
        traceback.print_exc()
        return None

def create_combined_excel(invoices_data, output_path):
    """Create combined Excel file - memory efficient version"""
    # Create workbook directly without pandas to save memory
    wb = Workbook()
    ws = wb.active
    ws.title = '所有发票'
    
    # Add header row
    ws.append(['品名', '数量', '单价', '金额'])
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add data from all invoices
    for i, invoice in enumerate(invoices_data):
        data = invoice['data']
        
        # Add invoice separator
        date = data.get('date', invoice['filename'])
        ws.append([f"=== 发票 {i+1} ===", '', '', date])
        ws.append(['', '', '', ''])
        
        # Add items
        if 'items' in data and data['items']:
            for item in data['items']:
                ws.append([
                    item.get('品名', ''),
                    item.get('数量', ''),
                    item.get('单价', ''),
                    item.get('金额', '')
                ])
        
        # Add separator
        ws.append(['', '', '', ''])
        ws.append(['=' * 60, '', '', ''])
        ws.append(['', '', '', ''])
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    
    # Center align all cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
    
    # Save
    wb.save(output_path)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    if 'files[]' not in request.files:
        flash('没有选择文件 / No files selected', 'error')
        return redirect(url_for('index'))
    
    files = request.files.getlist('files[]')
    
    if not files or files[0].filename == '':
        flash('没有选择文件 / No files selected', 'error')
        return redirect(url_for('index'))
    
    # Limit number of files
    max_files = app.config.get('MAX_FILES', 10)
    if len(files) > max_files:
        flash(f'最多上传{max_files}个文件 / Maximum {max_files} files allowed', 'error')
        return redirect(url_for('index'))
    
    # Create temporary directory for processing
    temp_dir = tempfile.mkdtemp()
    all_invoices = []
    
    try:
        import time
        start_time = time.time()
        
        # Process files ONE BY ONE to save memory
        for idx, file in enumerate(files, 1):
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filepath = os.path.join(temp_dir, filename)
                
                file_start = time.time()
                print(f"[{idx}/{len(files)}] Processing: {filename}")
                
                # Save file
                file.save(filepath)
                
                # Extract data from invoice
                data = extract_invoice_data(filepath)
                
                if data:
                    all_invoices.append({
                        'filename': filename,
                        'data': data,
                        'timestamp': datetime.now().isoformat()
                    })
                    file_time = time.time() - file_start
                    print(f"[{idx}/{len(files)}] ✓ Completed in {file_time:.1f}s")
                else:
                    print(f"[{idx}/{len(files)}] ✗ Failed to extract data")
                
                # Delete uploaded file IMMEDIATELY to free memory
                try:
                    os.remove(filepath)
                except:
                    pass
                
                # Force garbage collection after each file
                import gc
                gc.collect()
        
        total_time = time.time() - start_time
        print(f"Total processing time: {total_time:.1f}s for {len(files)} file(s)")
        
        if not all_invoices:
            flash('无法处理任何发票 / Could not process any invoices', 'error')
            return redirect(url_for('index'))
        
        # Create Excel file
        output_filename = f"所有发票_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(temp_dir, output_filename)
        create_combined_excel(all_invoices, output_path)
        
        # Send file to user
        response = send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Clean up temp directory after sending
        try:
            shutil.rmtree(temp_dir)
        except:
            pass
            
        return response
    
    except Exception as e:
        print(f"Error: {str(e)}")
        flash(f'处理出错 / Error: {str(e)}', 'error')
        # Clean up on error
        try:
            shutil.rmtree(temp_dir)
        except:
            pass
        return redirect(url_for('index'))

@app.route('/health')
def health():
    return {'status': 'ok', 'message': 'Invoice Scanner API is running'}

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)

